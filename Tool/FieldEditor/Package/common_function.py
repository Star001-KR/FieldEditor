from Package.Enum_Common import *
import openpyxl

def Get_ColumnProperty(_checkData):
    if str(_checkData).startswith("#"):
        return EColumnProperty.DesignColumn

    elif str(_checkData).endswith("?"):
        return EColumnProperty.NullableColumn

    elif str(_checkData).endswith("!"):
        return EColumnProperty.UniqueColumn

    else:
        return EColumnProperty.NormalColumn


def Copy_ExcelDataToLib(_fileFullName, _targetSheetName):
    dataDic = {}
    keyList = []
    
    EXCEL_FILE = openpyxl.load_workbook(_fileFullName)

    for sheetName in EXCEL_FILE.sheetnames:
        if sheetName == _targetSheetName:
            for row in EXCEL_FILE[sheetName].rows:
                if row[0].value == "None":
                    continue
                
                else:
                    dataDic[row[0].value] = row
                    keyList.append(row[0].value)

    EXCEL_FILE.close()

    if None in keyList:
        del dataDic[None]
        
        while(True):
            if None in keyList:
                keyList.pop()

            else:
                break
    
    return (dataDic, keyList)


def Copy_LibToExcelData(_fileFullName, _dataDic, _targetSheetName):
    EXCEL_FILE = openpyxl.load_workbook(_fileFullName)
    
    # clear excel data.
    for sheetName in EXCEL_FILE.sheetnames:
        if sheetName == _targetSheetName:
            for row in EXCEL_FILE[sheetName].rows:
                col = 0
                for _ in EXCEL_FILE[sheetName].columns:
                    row[col].value = None
                    col += 1
    
    # input excel data.
    for sheetName in EXCEL_FILE.sheetnames:
        if sheetName == _targetSheetName:
            for row in _dataDic.keys():
                for col in range(0, len(_dataDic[row])):
                    EXCEL_FILE[_targetSheetName].cell(row + 1, col + 1).value = _dataDic[row][col]

    EXCEL_FILE.save(_fileFullName)
    EXCEL_FILE.close()